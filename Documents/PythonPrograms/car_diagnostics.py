#Engine Link Car Dignostics Analysis Program
#Purpose: To take the log file created by the Engine Link App
# and put it into an easy to read Excel Spreadsheet for analysis
#Author Nate Epply
#Start Date 20SEP14

#Importing openpyxl module
from openpyxl import *

#Loading the Excel workbook that I am going to write the information to.
wb = Workbook()

#Asking for source file aka the engine log
file_choice = input("Which file would you like to import? ")

#Function to create and label sheets in the spreadsheet
def sheet(label):
    ws = wb.create_sheet()
    ws.title = label
    return ws


#Iterating through the engine log and writing to the excel spreadsheet
with open(file_choice,'r',encoding = 'utf-8') as my_file:
    SHEETS = ['Speed',' Acceleration',' Engine Power',' Instantaneous Fuel Effeciency',
              ' Average Fuel Effeciency',' Instantaneous MPG',' Average MPG',' MAF air flow rate',
              ' Accelerator pedal position E',' Commanded throttle actuator',' Engine Load',' Coolant Temp',
              ' RPM',' Torque',' Fuel level',' Intake air temperature',' Calculated engine load value',' Engine coolant temperature',
              ' Short term fuel % trim Bk 1',' Throttle position', 'Fuel Level Input','Catalyst Temp Bk 1 Sensor 1',
              ' Ambient air temperature',' Relative throttle position',' Timing advance',' Vehicle speed']
    for label in SHEETS:
        sheet(label)
        ws = wb.get_sheet_by_name(label)
        r = 1
        c = 1
        for line in my_file:
            if label in line:
                line = line.strip().split(',')
                for each in line:
                    if each == line[2]:
                        each = float(each)
                        ws.cell(row = r, column = c).value = each
                        c += 1
                    else:
                        ws.cell(row = r, column = c).value = each
                        c += 1
                    if each == line[3]:
                        r += 1
                        c = 1
        my_file.seek(0)

#CHARTING
from openpyxl.charts import *

#Asking user what sheet to pull the data from
start_chart = (input('Do you want to add a chart? ').lower())
YES = ['yes','y']

#Function for determining the last row with data in the sheet
def last_row(sheet):
    ws = wb.get_sheet_by_name(sheet)
    return len(ws.rows)

#Function that builds and adds the chart to the sheet
def build_chart(sheet,chart_type):
    ws = wb.get_sheet_by_name(sheet)
    values = Reference(ws, (1,3),(last_row(sheet),3))
    ws = wb.get_sheet_by_name('Sheet')
    series = Series(values, title = sheet)
    chart = charting_type(chart_type)
    chart.append(series)
    ws.add_chart(chart)

#Function for determining the type of chart to be created
#Currently the scatter chart is not working, all others are working
def charting_type(type_chart):
    if type_chart == 'barchart':
        return BarChart()
    elif type_chart == 'linechart':
        return LineChart()
    elif type_chart == 'scatterchart':
        return ScatterChart()
    elif type_chart == 'piechart':
        return PieChart()
    
#While loop to if want to create more than one chart
while start_chart in YES:
    data_set = input('Enter the name of the data you would like to chart? ')
    type_chart = (input('Enter the type of chart you would like to create? ').lower())
    build_chart(data_set,type_chart)
    start_chart = input('Do you want to build another chart? ')

#Saving the workbook and ending the program
wb.save('engine_diagnostics.xlsx')


